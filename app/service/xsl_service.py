import random

from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment


class ObjectiveDealValue(object):
    def __init__(self):
        self.STRATEGIC_PLAN = 150
        self.PERSPECTIVE_PLAN = 320
        self.OPERATIVE_PLAN = 500

    @property
    def strategic(self):
        return self.STRATEGIC_PLAN

    @property
    def perspective(self):
        return self.PERSPECTIVE_PLAN

    @property
    def operative(self):
        return self.OPERATIVE_PLAN


class Service(object):
    def __init__(self, file_path):
        self.file_path = file_path

        self.common_start_column = 5
        self.strategic_start_row = 1
        self.perspective_start_row = 5
        self.operative_start_row = 9

        self.common_indent = 16

    def write_sheet(self, data: dict):
        self.common_start_column = 5
        self.strategic_start_row = 1
        self.perspective_start_row = 5
        self.operative_start_row = 9

        self.common_indent = 16

        wb = Workbook()

        ws1 = wb.active
        ws1.title = 'Contact Details'

        ws1['A1'] = 'Contact Information'
        ws1.merge_cells('A1:B1')

        ws1['D1'] = 'Manager Information'
        ws1.merge_cells('D1:E1')

        ws1['G1'] = 'Extra Information'
        ws1.merge_cells('G1:H1')

        for i, (key, value) in enumerate(data['contactDetails'].items()):
            ws1.cell(row=(i + 2), column=1).value = key
            ws1.cell(row=(i + 2), column=2).value = value

        for i, (key, value) in enumerate(data['managerDetails'].items()):
            ws1.cell(row=(i + 2), column=4).value = key
            ws1.cell(row=(i + 2), column=5).value = value

        for i, (key, value) in enumerate(data['extraDetails'].items()):
            ws1.cell(row=(i + 2), column=7).value = key
            ws1.cell(row=(i + 2), column=8).value = value

        ws2 = wb.create_sheet(title="Products Information")
        start_indent = 1
        end_indent = 9

        for i, product in enumerate(data['products']):
            ws2.cell(column=1, row=start_indent, value=str(i + 1)).alignment = Alignment(horizontal='center',
                                                                                         vertical='center')
            ws2.merge_cells('A{}:A{}'.format(start_indent, start_indent + end_indent))

            for j, (key, value) in enumerate(product['details'].items()):
                ws2.cell(row=(j + start_indent), column=2).value = key
                ws2.cell(row=(j + start_indent), column=3).value = value

            self._write_strategic_plan(ws2, product['plans']['strategic'])
            self._write_perspective_plan(ws2, product['plans']['perspective'])
            self._write_operative_plan(ws2, product['plans']['operative'])

            start_indent += self.common_indent

        wb.save(filename=self.file_path)

    def calc_making_deal_probability(self, binary_file):
        wb = load_workbook(filename=BytesIO(binary_file.read()))
        ws = wb['Products Information']
        wa = wb['Products Information']
        wf = wb['Products Information']

        total_plans_result = self._get_plans_total_results(ws)
        result = self._update_with_probability(total_plans_result)

        odv = ObjectiveDealValue()
        result['objective_deal_values'] = {
            'strategic': odv.strategic,
            'perspective': odv.perspective,
            'operative': odv.operative
        }

        return result

    def _update_with_probability(self, products):
        result = {}
        odv = ObjectiveDealValue()

        products_count = len(products)
        meet_plans = 0

        for product in products:
            incr = 0
            if product['strategic'] >= odv.strategic:
                incr += 1
            if product['perspective'] >= odv.perspective:
                incr += 1
            if product['operative'] >= odv.operative:
                incr += 1
            if incr == 3:
                meet_plans += 1
                product['meet'] = True
            else:
                product['meet'] = False

        result['probability'] = 0 if meet_plans == 0 else (meet_plans / products_count) * 100
        result['meet_plans_count'] = meet_plans
        result['not_meet_plans_count'] = products_count - meet_plans
        result['products'] = products

        return result

    def _get_plans_total_results(self, ws):
        products = []

        product_number = 1
        indent = self.common_indent
        current_product_number_row = 1

        strategic_start_row = self.strategic_start_row + 1
        perspective_start_row = self.perspective_start_row + 1
        opearive_start_row = self.operative_start_row + 1

        while True:
            try:
                product = {}
                if ws['A{}'.format(current_product_number_row)].value == str(product_number):
                    product['name'] = ws['C{}'.format(current_product_number_row)].value
                    product['sku'] = ws['C{}'.format(current_product_number_row + 7)].value
                    for start_row, plan in zip((strategic_start_row, perspective_start_row, opearive_start_row),
                                               ('strategic', 'perspective', 'operative')):
                        for row in ws.iter_rows(min_row=start_row,
                                                max_row=start_row,
                                                min_col=self.common_start_column):
                            for cell in row:
                                if cell.value == 'Total':
                                    product[plan] = ws.cell(row=cell.row + 1, column=cell.col_idx).value
                                    break
                    products.append(product)
                else:
                    break
            except Exception as e:
                error = 'Method "_parse_plans_total_results". ' + repr(e)
                print(error)

            product_number += 1
            current_product_number_row += indent
            strategic_start_row += indent
            perspective_start_row += indent
            opearive_start_row += indent

        return products

    def _write_strategic_plan(self, ws, data):
        header_row = self.strategic_start_row
        for year, quarters_data in sorted(data.items()):
            color = self.random_color()
            if year == 'total':
                continue
            start_col = self.common_start_column
            start_row = self.strategic_start_row + 1

            cols_to_split = len(quarters_data) - 1

            ws.cell(row=start_row, column=start_col, value=year).alignment = Alignment(horizontal='center')

            ws.merge_cells(start_row=start_row,
                           end_row=start_row,
                           start_column=start_col,
                           end_column=start_col + cols_to_split)

            start_quarter_col = start_col
            for quarter, value in sorted(quarters_data.items()) if isinstance(quarters_data, dict) else []:
                ws.cell(row=start_row + 1,
                        column=start_quarter_col,
                        value='{} quarter'.format(quarter)).alignment = Alignment(horizontal='center')

                ws.cell(row=start_row + 2,
                        column=start_quarter_col,
                        value=value).alignment = Alignment(horizontal='center')

                start_quarter_col += 1

            self.common_start_column += cols_to_split + 1

            ws.cell(row=start_row, column=start_col).fill = PatternFill(fill_type="solid",
                                                                        start_color='FF' + color,
                                                                        end_color='FF' + color)

        self._write_sum(ws=ws,
                        header_row_number=self.strategic_start_row + 1,
                        col_number=self.common_start_column,
                        cols_to_split=2,
                        value=data['total'])

        ws.cell(row=header_row,
                column=5,
                value='Strategic Plan').alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=header_row,
                       end_row=header_row,
                       start_column=5,
                       end_column=self.common_start_column)

            # for row in ws.iter_rows(min_col=start_col,
            #                         max_col=start_col + cols_to_split,
            #                         min_row=start_row,
            #                         max_row=start_row + 2):
            #     for cell in row:
            #         cell.fill = PatternFill(fill_type="solid", start_color='FF' + color, end_color='FF' + color)

        self.strategic_start_row += self.common_indent
        self.common_start_column = 5

    def _write_perspective_plan(self, ws, data):
        header_row = self.perspective_start_row
        for year, months_data in sorted(data.items()):
            color = self.random_color()
            if year == 'total':
                continue
            start_row = self.perspective_start_row + 1
            start_col = self.common_start_column
            months_count = len(months_data) - 1

            ws.cell(row=start_row, column=start_col, value=year).alignment = Alignment(horizontal='center')
            ws.merge_cells(start_row=start_row,
                           end_row=start_row,
                           start_column=start_col,
                           end_column=start_col + months_count)

            start_month_col = start_col
            for month in Calendar.ordered_months(months_data):
                ws.cell(row=start_row + 1,
                        column=start_month_col,
                        value=month).alignment = Alignment(horizontal='center')

                ws.cell(row=start_row + 2,
                        column=start_month_col,
                        value=months_data[month]).alignment = Alignment(horizontal='center')

                start_month_col += 1

            ws.cell(row=start_row, column=start_col).fill = PatternFill(fill_type="solid",
                                                                        start_color='FF' + color,
                                                                        end_color='FF' + color)

            self.common_start_column += months_count + 1

        self._write_sum(ws=ws,
                        header_row_number=self.perspective_start_row + 1,
                        col_number=self.common_start_column,
                        cols_to_split=2,
                        value=data['total'])

        ws.cell(row=header_row,
                column=5,
                value='Perspective Plan').alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=header_row,
                       end_row=header_row,
                       start_column=5,
                       end_column=self.common_start_column)

        self.perspective_start_row += self.common_indent
        self.common_start_column = 5

    def _write_operative_plan(self, ws, data):
        month_start_col = self.common_start_column
        year_start_col = self.common_start_column
        header_row = self.operative_start_row

        for year, year_data in sorted(data.items()):
            color = self.random_color()
            if year == 'total':
                continue

            start_row = self.operative_start_row + 1

            decades_count = len(year_data) * 3 - 1

            ws.cell(row=start_row, column=year_start_col, value=year).alignment = Alignment(horizontal='center')
            ws.merge_cells(start_row=start_row,
                           end_row=start_row,
                           start_column=year_start_col,
                           end_column=year_start_col + decades_count)

            for month in Calendar.ordered_months(year_data):
                ws.cell(row=start_row + 1,
                        column=month_start_col,
                        value=month).alignment = Alignment(horizontal='center')

                ws.merge_cells(start_row=start_row + 1,
                               end_row=start_row + 1,
                               start_column=month_start_col,
                               end_column=month_start_col + 2)

                start_decade_col = month_start_col
                for decade_key, value in year_data[month].items():
                    ws.cell(row=start_row + 2,
                            column=start_decade_col,
                            value=decade_key).alignment = Alignment(horizontal='center')

                    ws.cell(row=start_row + 3,
                            column=start_decade_col,
                            value=value).alignment = Alignment(horizontal='center')

                    start_decade_col += 1

                month_start_col += 3

            ws.cell(row=start_row, column=year_start_col).fill = PatternFill(fill_type="solid",
                                                                             start_color='FF' + color,
                                                                             end_color='FF' + color)

            year_start_col += decades_count + 1

        self._write_sum(ws=ws,
                        header_row_number=self.operative_start_row + 1,
                        col_number=year_start_col,
                        cols_to_split=3,
                        value=data['total'])

        ws.cell(row=header_row,
                column=5,
                value='Operative Plan').alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=header_row,
                       end_row=header_row,
                       start_column=5,
                       end_column=month_start_col)

        self.operative_start_row += self.common_indent
        self.common_start_column = 5

    def _write_sum(self, ws, header_row_number, col_number, cols_to_split, value):
        ws.cell(row=header_row_number, column=col_number, value='Total').alignment = Alignment(horizontal='center')

        ws.cell(row=header_row_number, column=col_number).fill = PatternFill(fill_type="solid",
                                                                             start_color='FF4500',
                                                                             end_color='FF4500')

        ws.cell(row=header_row_number + 1, column=col_number, value=value).alignment = Alignment(horizontal='center',
                                                                                                 vertical='center')
        ws.cell(row=header_row_number + 1, column=col_number).fill = PatternFill(fill_type="solid",
                                                                                 start_color='B0F2F4',
                                                                                 end_color='B0F2F4')
        ws.merge_cells(start_row=header_row_number + 1,
                       end_row=header_row_number + cols_to_split,
                       start_column=col_number,
                       end_column=col_number)

    def random_color(self):
        r = lambda: random.randint(0, 255)
        return '%02X%02X%02X' % (r(), r(), r())


class Calendar(object):

    @staticmethod
    def months_map():
        return (
            'January',
            'February',
            'March',
            'April',
            'May',
            'June',
            'July',
            'August',
            'September',
            'November',
            'October',
            'December',
        )

    @staticmethod
    def ordered_months(months_dict):
        return [month for month in Calendar.months_map() if month in months_dict]
