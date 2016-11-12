from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
import smtplib
from os.path import basename
import mimetypes
from email.mime.multipart import MIMEMultipart
from email import encoders
from email.message import Message
from email.mime.audio import MIMEAudio
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.text import MIMEText
from email.utils import COMMASPACE, formatdate
import random
import os


class Service(object):
    def __init__(self, file_path):
        self.file_path = file_path
        self.sender = 'alex.chudovsky@gmail.com'
        self.username = 'alex.chudovsky@gmail.com'
        self.pwd = 'PWD'

        self.common_start_column = 5
        self.strategic_start_row = 1
        self.perspective_start_row = 5
        self.operative_start_row = 9

    def send_email(self, recipient_email, recipient_name, subject, data):

        self.write_sheet(data)

        msg = MIMEMultipart()
        msg['From'] = self.sender
        msg['To'] = recipient_email
        msg['Date'] = formatdate(localtime=True)
        msg['Subject'] = subject

        msg.attach(MIMEText('Dear {}. Please, give a little time to review our suggestion.\n'
                            'Thanks for watching!'.format(recipient_name)))

        part = MIMEBase('application', "octet-stream")
        part.set_payload(open(self.file_path, "rb").read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment; filename="{0}"'.format(os.path.basename(self.file_path)))
        msg.attach(part)

        server, port = 'smtp.gmail.com', 587
        smtp = smtplib.SMTP(server, port)
        smtp.starttls()
        smtp.login(self.username, self.pwd)
        smtp.sendmail(self.sender, recipient_email, msg.as_string())
        smtp.quit()

    def write_sheet(self, data: dict):
        wb = Workbook()

        ws1 = wb.active
        ws1.title = 'Contact Details'

        ws1['A1'] = 'Contact Information'
        ws1.merge_cells('A1:B1')

        ws1['D1'] = 'Manager Information'
        ws1.merge_cells('D1:E1')

        ws1['G1'] = 'Extra Information'
        ws1.merge_cells('G1:H1')

        for i, (key, value) in enumerate(data['contactDetails'].items()):
            ws1.cell(row=(i + 2), column=1).value = key
            ws1.cell(row=(i + 2), column=2).value = value

        for i, (key, value) in enumerate(data['managerDetails'].items()):
            ws1.cell(row=(i + 2), column=4).value = key
            ws1.cell(row=(i + 2), column=5).value = value

        for i, (key, value) in enumerate(data['extraDetails'].items()):
            ws1.cell(row=(i + 2), column=7).value = key
            ws1.cell(row=(i + 2), column=8).value = value

        ws2 = wb.create_sheet(title="Products Information")
        start_indent = 1
        end_indent = 9

        for i, product in enumerate(data['products']):
            ws2.cell(column=1, row=start_indent, value=str(i + 1)).alignment = Alignment(horizontal='center',
                                                                                         vertical='center')
            ws2.merge_cells('A{}:A{}'.format(start_indent, start_indent + end_indent))

            for j, (key, value) in enumerate(product['details'].items()):
                ws2.cell(row=(j + start_indent), column=2).value = key
                ws2.cell(row=(j + start_indent), column=3).value = value

            self._write_strategic_plan(ws2, product['plans']['strategic'])
            self._write_perspective_plan(ws2, product['plans']['perspective'])
            self._write_operative_plan(ws2, product['plans']['operative'])

            start_indent += 16

        wb.save(filename=self.file_path)

    def _write_strategic_plan(self, ws, data):
        header_row = self.strategic_start_row
        for year, quarters_data in sorted(data.items()):
            color = self.random_color()
            if year == 'total':
                continue
            start_col = self.common_start_column
            start_row = self.strategic_start_row + 1

            cols_to_split = len(quarters_data) - 1

            ws.cell(row=start_row, column=start_col, value=year).alignment = Alignment(horizontal='center')

            ws.merge_cells(start_row=start_row,
                           end_row=start_row,
                           start_column=start_col,
                           end_column=start_col + cols_to_split)

            start_quarter_col = start_col
            for quarter, value in sorted(quarters_data.items()) if isinstance(quarters_data, dict) else []:
                ws.cell(row=start_row + 1,
                        column=start_quarter_col,
                        value='{} quarter'.format(quarter)).alignment = Alignment(horizontal='center')

                ws.cell(row=start_row + 2,
                        column=start_quarter_col,
                        value=value).alignment = Alignment(horizontal='center')

                start_quarter_col += 1

            self.common_start_column += cols_to_split + 1

            ws.cell(row=start_row, column=start_col).fill = PatternFill(fill_type="solid",
                                                                        start_color='FF' + color,
                                                                        end_color='FF' + color)

        self._write_sum(ws=ws,
                        header_row_number=self.strategic_start_row + 1,
                        col_number=self.common_start_column,
                        cols_to_split=2,
                        value=data['total'])

        ws.cell(row=header_row,
                column=5,
                value='Strategic Plan').alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=header_row,
                       end_row=header_row,
                       start_column=5,
                       end_column=self.common_start_column)

            # for row in ws.iter_rows(min_col=start_col,
            #                         max_col=start_col + cols_to_split,
            #                         min_row=start_row,
            #                         max_row=start_row + 2):
            #     for cell in row:
            #         cell.fill = PatternFill(fill_type="solid", start_color='FF' + color, end_color='FF' + color)

        self.strategic_start_row += 16
        self.common_start_column = 5

    def _write_perspective_plan(self, ws, data):
        header_row = self.perspective_start_row
        for year, months_data in sorted(data.items()):
            color = self.random_color()
            if year == 'total':
                continue
            start_row = self.perspective_start_row + 1
            start_col = self.common_start_column
            months_count = len(months_data) - 1

            ws.cell(row=start_row, column=start_col, value=year).alignment = Alignment(horizontal='center')
            ws.merge_cells(start_row=start_row,
                           end_row=start_row,
                           start_column=start_col,
                           end_column=start_col + months_count)

            start_month_col = start_col
            for month in Calendar.ordered_months(months_data):
                ws.cell(row=start_row + 1,
                        column=start_month_col,
                        value=month).alignment = Alignment(horizontal='center')

                ws.cell(row=start_row + 2,
                        column=start_month_col,
                        value=months_data[month]).alignment = Alignment(horizontal='center')

                start_month_col += 1

            ws.cell(row=start_row, column=start_col).fill = PatternFill(fill_type="solid",
                                                                        start_color='FF' + color,
                                                                        end_color='FF' + color)

            self.common_start_column += months_count + 1

        self._write_sum(ws=ws,
                        header_row_number=self.perspective_start_row + 1,
                        col_number=self.common_start_column,
                        cols_to_split=2,
                        value=data['total'])

        ws.cell(row=header_row,
                column=5,
                value='Perspective Plan').alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=header_row,
                       end_row=header_row,
                       start_column=5,
                       end_column=self.common_start_column)

        self.perspective_start_row += 16
        self.common_start_column = 5

    def _write_operative_plan(self, ws, data):
        month_start_col = self.common_start_column
        year_start_col = self.common_start_column
        header_row = self.operative_start_row

        for year, year_data in sorted(data.items()):
            color = self.random_color()
            if year == 'total':
                continue

            start_row = self.operative_start_row + 1

            decades_count = len(year_data) * 3 - 1

            ws.cell(row=start_row, column=year_start_col, value=year).alignment = Alignment(horizontal='center')
            ws.merge_cells(start_row=start_row,
                           end_row=start_row,
                           start_column=year_start_col,
                           end_column=year_start_col + decades_count)

            for month in Calendar.ordered_months(year_data):
                ws.cell(row=start_row + 1,
                        column=month_start_col,
                        value=month).alignment = Alignment(horizontal='center')

                ws.merge_cells(start_row=start_row + 1,
                               end_row=start_row + 1,
                               start_column=month_start_col,
                               end_column=month_start_col + 2)

                start_decade_col = month_start_col
                for decade_key, value in year_data[month].items():
                    ws.cell(row=start_row + 2,
                            column=start_decade_col,
                            value=decade_key).alignment = Alignment(horizontal='center')

                    ws.cell(row=start_row + 3,
                            column=start_decade_col,
                            value=value).alignment = Alignment(horizontal='center')

                    start_decade_col += 1

                month_start_col += 3

            ws.cell(row=start_row, column=year_start_col).fill = PatternFill(fill_type="solid",
                                                                             start_color='FF' + color,
                                                                             end_color='FF' + color)

            year_start_col += decades_count + 1

        self._write_sum(ws=ws,
                        header_row_number=self.operative_start_row + 1,
                        col_number=year_start_col,
                        cols_to_split=3,
                        value=data['total'])

        ws.cell(row=header_row,
                column=5,
                value='Operative Plan').alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=header_row,
                       end_row=header_row,
                       start_column=5,
                       end_column=month_start_col)

        self.operative_start_row += 16
        self.common_start_column = 5

    def _write_sum(self, ws, header_row_number, col_number, cols_to_split, value):
        ws.cell(row=header_row_number, column=col_number, value='Total').alignment = Alignment(horizontal='center')

        ws.cell(row=header_row_number, column=col_number).fill = PatternFill(fill_type="solid",
                                                                             start_color='FF4500',
                                                                             end_color='FF4500')

        ws.cell(row=header_row_number + 1, column=col_number, value=value).alignment = Alignment(horizontal='center',
                                                                                                 vertical='center')
        ws.cell(row=header_row_number + 1, column=col_number).fill = PatternFill(fill_type="solid",
                                                                                 start_color='B0F2F4',
                                                                                 end_color='B0F2F4')
        ws.merge_cells(start_row=header_row_number + 1,
                       end_row=header_row_number + cols_to_split,
                       start_column=col_number,
                       end_column=col_number)

    def random_color(self):
        r = lambda: random.randint(0, 255)
        return '%02X%02X%02X' % (r(), r(), r())


class Calendar(object):

    @staticmethod
    def months_map():
        return (
            'January',
            'February',
            'March',
            'April',
            'May',
            'June',
            'July',
            'August',
            'September',
            'November',
            'October',
            'December',
        )

    @staticmethod
    def ordered_months(months_dict):
        return [month for month in Calendar.months_map() if month in months_dict]
